import json
import base64
import binascii
from io import BytesIO, TextIOWrapper
import csv
from openpyxl import Workbook, load_workbook

def lambda_handler(event, context):
    # 1. Normalize headers
    headers = {k.lower(): v for k, v in (event.get('headers') or {}).items()}
    content_type = headers.get('content-type', '')

    # 2. Determine conversion mode
    if 'text/csv' in content_type:
        mode = 'csv_to_xlsx'
    elif 'spreadsheetml' in content_type or 'xlsx' in content_type:
        mode = 'xlsx_to_csv'
    else:
        return {
            'statusCode': 400,
            'body': json.dumps({
                'error': 'Unsupported Content-Type; must be text/csv or application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })
        }

    # 3. Decode body
    raw = event.get('body') or ''
    if event.get('isBase64Encoded', False):
        try:
            payload = base64.b64decode(raw)
        except binascii.Error as err:
            return {
                'statusCode': 400,
                'body': json.dumps({'error': 'Invalid base64 payload'})
            }
    else:
        payload = raw.encode('utf-8')

    buf_in = BytesIO(payload)
    buf_out = BytesIO()

    try:
        if mode == 'csv_to_xlsx':
            # CSV → XLSX
            reader = csv.reader(TextIOWrapper(buf_in, encoding='utf-8'))
            wb = Workbook()
            ws = wb.active
            for row in reader:
                ws.append(row)
            wb.save(buf_out)
            output_content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = 'output.xlsx'
        else:
            # XLSX → CSV
            wb = load_workbook(buf_in, read_only=True)
            ws = wb.active
            writer = csv.writer(TextIOWrapper(buf_out, encoding='utf-8', newline=''))
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
            output_content_type = 'text/csv'
            filename = 'output.csv'

        # 4. Base64-encode response
        data = buf_out.getvalue()
        body_b64 = base64.b64encode(data).decode('utf-8')

        return {
            'statusCode': 200,
            'isBase64Encoded': True,
            'headers': {
                'Content-Type': output_content_type,
                'Content-Disposition': f'attachment; filename="{filename}"'
            },
            'body': body_b64
        }

    except Exception as e:
        # In practice also log traceback to CloudWatch here
        return {
            'statusCode': 500,
            'body': json.dumps({'error': str(e)})
        }

